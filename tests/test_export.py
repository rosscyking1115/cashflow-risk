"""Behaviour of the Excel (.xlsx) export — the accountant-facing deliverable.

A tenant-scoped, authenticated download rendered from the persisted analysis,
with CSV/formula-injection escaping (an accountant opens these in Excel).
"""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from cashflow_risk.api import app
from cashflow_risk.auth import Principal, mint_token
from cashflow_risk.auth.settings import jwt_secret

client = TestClient(app)

HEADER = "invoice_id,customer_id,amount,issue_date,due_date\n"


def _auth(business_id: str) -> dict[str, str]:
    token = mint_token(
        Principal(user_id="u1", business_id=business_id, email="u1@example.co"),
        secret=jwt_secret(),
    )
    return {"Authorization": f"Bearer {token}"}


def _upload(csv_text: str, business_id: str):
    return client.post(
        "/api/analyze",
        files={"invoices": ("inv.csv", csv_text, "text/csv")},
        data={"opening_balance": "5000", "minimum_reserve": "2000"},
        headers=_auth(business_id),
    )


def test_export_returns_an_xlsx_with_the_expected_sheets() -> None:
    run_id = _upload(HEADER + "INV-1,Acme Ltd,5000,2026-01-01,2026-01-31\n", "biz-x").json()[
        "run_id"
    ]

    resp = client.get(f"/api/runs/{run_id}/export.xlsx", headers=_auth("biz-x"))

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content))
    assert {"Summary", "Forecast", "Cash at risk"} <= set(wb.sheetnames)
    assert wb["Forecast"].max_row == 14  # header + 13 weeks


def test_export_escapes_formula_injection_in_text_cells() -> None:
    # a malicious customer name that Excel would otherwise execute as a formula
    csv_text = HEADER + "INV-1,=HYPERLINK(evil),5000,2026-01-01,2026-01-31\n"
    run_id = _upload(csv_text, "biz-x").json()["run_id"]

    resp = client.get(f"/api/runs/{run_id}/export.xlsx", headers=_auth("biz-x"))
    wb = load_workbook(BytesIO(resp.content))
    customers = [row[1].value for row in wb["Cash at risk"].iter_rows(min_row=2)]

    assert any(isinstance(v, str) and v.startswith("'=") for v in customers)


def test_export_is_tenant_scoped() -> None:
    run_id = _upload(HEADER + "INV-1,Acme,5000,2026-01-01,2026-01-31\n", "tenant-a").json()[
        "run_id"
    ]

    url = f"/api/runs/{run_id}/export.xlsx"
    assert client.get(url, headers=_auth("tenant-b")).status_code == 404
    assert client.get(url, headers=_auth("tenant-a")).status_code == 200


def test_export_requires_authentication() -> None:
    assert client.get("/api/runs/anything/export.xlsx").status_code == 401
