"""Render a persisted analysis as an accountant-facing .xlsx workbook.

Three sheets — Summary, Forecast, Cash at risk. Text cells are escaped against
CSV/formula injection (a cell beginning ``= + - @`` is prefixed with an
apostrophe), because accountants open these exports in Excel/Sheets.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from cashflow_risk.api.schemas import AnalysisResponse

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MONEY = "£#,##0"
_HEADER_FONT = Font(bold=True)


def _safe(value: object) -> object:
    """Neutralise spreadsheet formula injection in text cells."""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _header(ws: Worksheet, labels: list[str]) -> None:
    ws.append(labels)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def analysis_to_xlsx(response: AnalysisResponse) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    for label, value in [
        ("Business", response.business_id),
        ("As of", response.as_of.isoformat()),
        ("Runway (weeks)", response.brief.runway_weeks),
        ("Shortfall", "Yes" if response.brief.has_shortfall else "No"),
        ("Minimum reserve", response.minimum_reserve),
        ("Headline", response.brief.headline),
        ("Disclaimer", response.brief.disclaimer),
    ]:
        summary.append([label, _safe(value)])

    forecast = wb.create_sheet("Forecast")
    _header(forecast, ["Week", "Week start", "Opening", "Inflow", "Outflow", "Closing"])
    for w in response.weeks:
        forecast.append(
            [
                w.index + 1,
                w.week_start.isoformat(),
                w.opening_balance,
                w.expected_inflow,
                w.expected_outflow,
                w.closing_balance,
            ]
        )
    for col in ("C", "D", "E", "F"):
        for cell in forecast[col][1:]:
            cell.number_format = _MONEY

    risk = wb.create_sheet("Cash at risk")
    # "Risk score", not "Probability" — the 0-1 output ranks invoices, it is not
    # calibrated (measured mean ECE 0.212). See cashflow_risk.risk.
    _header(risk, ["Invoice", "Customer", "Risk score", "Band", "Cash at risk", "Why"])
    for s in response.top_risks:
        risk.append(
            [
                _safe(s.invoice_id),
                _safe(s.customer_id),
                s.probability,
                s.band,
                s.cash_at_risk,
                _safe(s.drivers[0] if s.drivers else ""),
            ]
        )
    for cell in risk["C"][1:]:
        cell.number_format = "0%"
    for cell in risk["E"][1:]:
        cell.number_format = _MONEY

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
